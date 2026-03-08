from __future__ import annotations

import argparse
import csv
import json
import math
import re
import shutil
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font

try:
    import requests
except Exception:
    requests = None

# =========================
# CONFIG
# =========================
BASE_DIR = Path(__file__).resolve().parent          # scripts_completar_info
PROJECT_DIR = BASE_DIR.parent                       # Lorena
TABLAS_DIR = PROJECT_DIR / "tablas"

FILENAME = str(PROJECT_DIR / "BaseDatos.xlsx")

SHEET_MASTER = "Master Localizacion - Corregida"
SHEET_BD = "BD 202603"
SHEET_LOC = "Maestro Localidades - Corregido"
SHEET_REF = "ID localidad-provincia-region"
SHEET_LEGEND = "Leyenda_Colores"
SHEET_AUDIT = "Auditoria_Geografia"

CSV_JERARQUIA = str(TABLAS_DIR / "jerarquia.csv")
CSV_LOCALIDADES = str(TABLAS_DIR / "localidades.csv")
JSON_ZIP = str(TABLAS_DIR / "argentina_zip_codes.json")

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
USER_AGENT = "MasterLocalizacionCleaner/4.2"

# =========================
# COLORES
# =========================
FILL_FROM_BD = PatternFill(start_color="9FC5E8", end_color="9FC5E8", fill_type="solid")
FILL_FROM_LOC = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
FILL_FROM_JER = PatternFill(start_color="D9A2E9", end_color="D9A2E9", fill_type="solid")
FILL_FROM_LOCCSV = PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid")
FILL_FROM_JSON = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
FILL_FROM_WEB = PatternFill(start_color="EA9999", end_color="EA9999", fill_type="solid")

FILL_NORMALIZED = PatternFill(start_color="D0E0E3", end_color="D0E0E3", fill_type="solid")

FILL_CONFLICT_MASTER = PatternFill(start_color="E06666", end_color="E06666", fill_type="solid")
FILL_CONFLICT_REF = PatternFill(start_color="8E7CC3", end_color="8E7CC3", fill_type="solid")
FILL_CONFLICT_JER = PatternFill(start_color="C27BA0", end_color="C27BA0", fill_type="solid")
FILL_CONFLICT_LOCCSV = PatternFill(start_color="76A5AF", end_color="76A5AF", fill_type="solid")
FILL_CONFLICT_JSON = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")

FILL_HEADER = PatternFill(start_color="6FA8DC", end_color="6FA8DC", fill_type="solid")

# =========================
# CAMPOS
# =========================
MASTER_FIELDS = [
    "GLN",
    "region_pais_cod",
    "amba/interior",
    "region_cod",
    "region_descr",
    "zona_cod",
    "zona_descr",
    "cod_prov",
    "provincia",
    "cod_localidad",
    "localidad",
    "departamento",
    "cp",
    "direccion",
    "lat",
    "long",
]

GEO_FIELDS = [
    "region_pais_cod",
    "amba/interior",
    "region_cod",
    "region_descr",
    "zona_cod",
    "zona_descr",
    "cod_prov",
    "provincia",
    "departamento",
    "cod_localidad",
    "localidad",
    "cp",
]

BD_TO_MASTER = {
    "gln": "GLN",
    "region_pais_cod": "region_pais_cod",
    "amba/interior": "amba/interior",
    "region_cod_ccu": "region_cod",
    "region_descip_ccu": "region_descr",
    "zona_cod": "zona_cod",
    "zona_descr": "zona_descr",
    "cod_provinica": "cod_prov",
    "provincia": "provincia",
    "cod_localidad": "cod_localidad",
    "localidad": "localidad",
    "departamento_ccu": "departamento",
    "sucursal_codigo_postal": "cp",
    "Direccion": "direccion",
    "latitud": "lat",
    "longitud": "long",
}

LOC_TO_MASTER = {
    "cod_localidad": "cod_localidad",
    "localidad": "localidad",
    "departamento": "departamento",
    "cod_provincia": "cod_prov",
    "provincia": "provincia",
    "cp": "cp",
    "region_pais_cod": "region_pais_cod",
    "amba/interior": "amba/interior",
    "region_cod": "region_cod",
    "region_descr": "region_descr",
    "zona_cod": "zona_cod",
    "zona_descr": "zona_descr",
}

REF_TO_MASTER = {
    "region_pais_cod": "region_pais_cod",
    "amba/interior": "amba/interior",
    "region_cod": "region_cod",
    "region_descip": "region_descr",
    "zona_cod": "zona_cod",
    "zona_descr": "zona_descr",
    "cod_provinica": "cod_prov",
    "provincia": "provincia",
    "cod_localidad": "cod_localidad",
    "localidad": "localidad",
}

JER_TO_MASTER = {
    "cod_localidad": "cod_localidad",
    "localidad": "localidad",
    "cod_prov": "cod_prov",
    "provincia": "provincia",
    "zona_cod": "zona_cod",
    "zona_descr": "zona_descr",
    "region_cod": "region_cod",
    "region_descr": "region_descr",
    "amba_interior": "amba/interior",
    "region_pais_cod": "region_pais_cod",
    "departamento": "departamento",
    "cp": "cp",
}

LOCCSV_TO_MASTER = {
    "cp": "cp",
    "localidad": "localidad",
}

JSON_FIELDS = ["cp", "departamento"]

CODE_FIELDS = {"GLN", "region_pais_cod", "region_cod", "zona_cod", "cod_prov", "cod_localidad", "cp"}
TEXT_UPPER_FIELDS = {"amba/interior"}
TITLE_FIELDS = {"provincia", "localidad", "departamento", "region_descr", "zona_descr"}
LATLON_FIELDS = {"lat", "long"}

# =========================
# STATS
# =========================
@dataclass
class Stats:
    from_bd: int = 0
    from_loc: int = 0
    from_jer: int = 0
    from_loccsv: int = 0
    from_json: int = 0
    from_web: int = 0
    normalized: int = 0
    rows_touched: int = 0

    conflict_rows_ref: int = 0
    conflicts_ref: int = 0

    conflict_rows_jer: int = 0
    conflicts_jer: int = 0

    conflict_rows_loccsv: int = 0
    conflicts_loccsv: int = 0

    conflict_rows_json: int = 0
    conflicts_json: int = 0


# =========================
# HELPERS
# =========================
def strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", str(text)).strip()


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    return s == "" or s.lower() in {"none", "nan", "null", "faltante", "s/d", "sin dato", "0"}


def normalize_key(text: Any) -> str:
    if is_blank(text):
        return ""
    s = clean_spaces(text)
    s = strip_accents(s).upper()
    return s


def title_case(text: str) -> str:
    small = {"de", "del", "la", "las", "los", "y"}
    words = clean_spaces(text).lower().split(" ")
    out = []
    for i, w in enumerate(words):
        if i > 0 and w in small:
            out.append(w)
        elif w in {"gba", "caba", "noa", "nea"}:
            out.append(w.upper())
        elif w == "cap":
            out.append("CAP")
        else:
            out.append(w.capitalize())
    return " ".join(out)


def parse_int_like(value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return str(int(value))
        return str(value).replace(",", ".")
    s = clean_spaces(value)
    s = s.replace(".0", "") if re.fullmatch(r"\d+\.0", s) else s
    s = s.replace(",", "") if re.fullmatch(r"\d{1,3}(,\d{3})+", s) else s
    s = s.replace(" ", "")
    if re.fullmatch(r"\d+", s):
        return s
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def normalize_id_sucursal(value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    s = clean_spaces(value).lower().replace(" ", "")
    raw = s[4:] if s.startswith("suc_") else s
    raw = raw.replace(",", "")
    if re.fullmatch(r"\d+\.0+", raw):
        raw = raw.split(".")[0]
    if re.fullmatch(r"\d+", raw):
        return f"suc_{int(raw)}"
    return f"suc_{raw}"


def normalize_latlon(value: Any) -> Optional[float]:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    s = clean_spaces(value)
    if re.fullmatch(r"-?\d+,\d+", s):
        s = s.replace(",", ".")
    elif re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", s):
        s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def normalize_field_value(field: str, value: Any) -> Any:
    if is_blank(value):
        return None
    if field == "ID Sucursal":
        return normalize_id_sucursal(value)
    if field in CODE_FIELDS:
        return parse_int_like(value)
    if field in LATLON_FIELDS:
        return normalize_latlon(value)
    s = clean_spaces(value)
    if field in TEXT_UPPER_FIELDS:
        return strip_accents(s).upper()
    if field in TITLE_FIELDS:
        return title_case(strip_accents(s))
    return s


def comparable_value(field: str, value: Any) -> Optional[str]:
    v = normalize_field_value(field, value)
    if v is None:
        return None
    if isinstance(v, float):
        return f"{v:.8f}"
    return str(v)


def find_headers(ws) -> Dict[str, int]:
    return {
        str(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value is not None
    }


def recreate_sheet(wb, name: str):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        ws_old = wb[name]
        wb.remove(ws_old)
        ws_new = wb.create_sheet(name, idx)
    else:
        ws_new = wb.create_sheet(name)
    return ws_new


def set_if_missing(cell, value, fill) -> bool:
    if value is None:
        return False
    if is_blank(cell.value):
        cell.value = value
        cell.fill = fill
        return True
    return False


def normalize_row(ws, row_num: int, cols: Dict[str, int], stats: Stats) -> bool:
    touched = False
    for field in ["ID Sucursal"] + MASTER_FIELDS:
        if field not in cols:
            continue
        cell = ws.cell(row=row_num, column=cols[field])
        new_val = normalize_field_value(field, cell.value)
        old_cmp = comparable_value(field, cell.value)
        new_cmp = comparable_value(field, new_val)
        if old_cmp != new_cmp:
            cell.value = new_val
            if old_cmp is not None or new_cmp is not None:
                cell.fill = FILL_NORMALIZED
                stats.normalized += 1
                touched = True
    return touched


def make_backup(path: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    original = Path(path).resolve()

    if not original.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel: {original}")

    backup_dir = original.parent / "BackUps"
    backup_dir.mkdir(exist_ok=True)
    backup_name = original.stem + f"_backup_{timestamp}" + original.suffix
    backup_path = backup_dir / backup_name
    shutil.copy2(original, backup_path)
    return str(backup_path)


# =========================
# LECTURA EXCEL
# =========================
def build_bd_index(ws_bd, bd_cols: Dict[str, int]) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = {}
    for row in ws_bd.iter_rows(min_row=2, values_only=True):
        row_map = {h: row[idx - 1] for h, idx in bd_cols.items()}
        key = normalize_id_sucursal(row_map.get("ID Sucursal"))
        if not key:
            continue
        cleaned = {}
        for src, dst in BD_TO_MASTER.items():
            cleaned[dst] = normalize_field_value(dst, row_map.get(src))
        score = sum(v is not None for v in cleaned.values())
        prev = index.get(key)
        prev_score = prev.get("__score__", -1) if prev else -1
        if score > prev_score:
            cleaned["__score__"] = score
            index[key] = cleaned
    return index


def build_loc_indexes(ws_loc, loc_cols: Dict[str, int]) -> Tuple[Dict[str, Dict[str, Any]], Dict[Tuple[str, str], Dict[str, Any]]]:
    by_cod: Dict[str, Dict[str, Any]] = {}
    by_prov_loc: Dict[Tuple[str, str], Dict[str, Any]] = {}

    for row in ws_loc.iter_rows(min_row=2, values_only=True):
        row_map = {h: row[idx - 1] for h, idx in loc_cols.items() if h is not None}
        cleaned = {}
        for src, dst in LOC_TO_MASTER.items():
            if src in row_map:
                cleaned[dst] = normalize_field_value(dst, row_map.get(src))

        cod_loc = cleaned.get("cod_localidad")
        if cod_loc:
            by_cod[cod_loc] = cleaned

        prov = normalize_key(cleaned.get("provincia"))
        loc = normalize_key(cleaned.get("localidad"))
        if prov and loc:
            by_prov_loc[(prov, loc)] = cleaned

    return by_cod, by_prov_loc


def build_ref_index(ws_ref, ref_cols: Dict[str, int]) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = {}
    for row in ws_ref.iter_rows(min_row=2, values_only=True):
        row_map = {h: row[idx - 1] for h, idx in ref_cols.items()}
        key = normalize_id_sucursal(row_map.get("Id local"))
        if not key:
            continue
        cleaned = {}
        for src, dst in REF_TO_MASTER.items():
            if src in row_map:
                cleaned[dst] = normalize_field_value(dst, row_map.get(src))
        index[key] = cleaned
    return index


# =========================
# LECTURA CSV / JSON
# =========================
def sniff_delimiter(path: str) -> str:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
        return dialect.delimiter
    except Exception:
        return ";"


def read_csv_rows(path: str) -> List[Dict[str, str]]:
    if not Path(path).exists():
        return []
    delimiter = sniff_delimiter(path)
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return [dict(row) for row in reader]


def build_jerarquia_indexes(path: str) -> Tuple[Dict[str, Dict[str, Any]], Dict[Tuple[str, str], Dict[str, Any]]]:
    rows = read_csv_rows(path)
    by_cod: Dict[str, Dict[str, Any]] = {}
    by_prov_loc: Dict[Tuple[str, str], Dict[str, Any]] = {}

    for row in rows:
        cleaned: Dict[str, Any] = {}
        for src, dst in JER_TO_MASTER.items():
            cleaned[dst] = normalize_field_value(dst, row.get(src))

        cod_loc = cleaned.get("cod_localidad")
        if cod_loc:
            by_cod[cod_loc] = cleaned

        prov = normalize_key(cleaned.get("provincia"))
        loc = normalize_key(cleaned.get("localidad"))
        if prov and loc:
            by_prov_loc[(prov, loc)] = cleaned

    return by_cod, by_prov_loc


def build_localidades_csv_index(path: str) -> Dict[Tuple[str, str], Dict[str, Any]]:
    rows = read_csv_rows(path)
    by_cp_loc: Dict[Tuple[str, str], Dict[str, Any]] = {}

    for row in rows:
        cleaned: Dict[str, Any] = {}
        for src, dst in LOCCSV_TO_MASTER.items():
            cleaned[dst] = normalize_field_value(dst, row.get(src))

        cp = cleaned.get("cp")
        loc = normalize_key(cleaned.get("localidad"))
        if cp and loc:
            by_cp_loc[(cp, loc)] = {
                "cp": cleaned.get("cp"),
                "localidad": normalize_field_value("localidad", row.get("localidad")),
                "idProvincia": parse_int_like(row.get("idProvincia")),
            }

    return by_cp_loc


def load_json_zip_indexes(path: str) -> Tuple[
    Dict[Tuple[str, str], Dict[str, Any]],
    Dict[Tuple[str, str, str], Dict[str, Any]]
]:
    by_prov_loc: Dict[Tuple[str, str], Dict[str, Any]] = {}
    by_prov_dept_loc: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    if not Path(path).exists():
        return by_prov_loc, by_prov_dept_loc

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    for prov_raw, depts in data.items():
        prov = normalize_key(prov_raw)
        if not isinstance(depts, dict):
            continue

        for dept_raw, loc_parents in depts.items():
            dept_norm = normalize_field_value("departamento", dept_raw)
            dept_key = normalize_key(dept_raw)

            if not isinstance(loc_parents, dict):
                continue

            for _loc_parent_raw, sub_locs in loc_parents.items():
                if not isinstance(sub_locs, dict):
                    continue

                for sub_loc_raw, cp_raw in sub_locs.items():
                    loc_key = normalize_key(sub_loc_raw)
                    cp = parse_int_like(cp_raw)
                    if not loc_key:
                        continue

                    record = {"cp": cp, "departamento": dept_norm}
                    by_prov_loc.setdefault((prov, loc_key), record)
                    by_prov_dept_loc[(prov, dept_key, loc_key)] = record

    return by_prov_loc, by_prov_dept_loc


# =========================
# WEB
# =========================
def enrich_from_web(address: str, locality: str, province: str) -> Tuple[Optional[str], Optional[str]]:
    if requests is None:
        return None, None

    params = {
        "q": f"{address}, {locality}, {province}, Argentina",
        "format": "json",
        "addressdetails": 1,
        "limit": 1,
    }
    headers = {"User-Agent": USER_AGENT}

    try:
        r = requests.get(NOMINATIM_URL, params=params, headers=headers, timeout=15)
        if r.status_code != 200:
            return None, None
        data = r.json()
        if not data:
            return None, None

        addr = data[0].get("address", {})
        cp = addr.get("postcode")
        dept = addr.get("county") or addr.get("state_district") or addr.get("city_district")

        cp = parse_int_like(cp) if cp else None
        dept = title_case(strip_accents(dept)) if dept else None
        return cp, dept
    except Exception:
        return None, None


# =========================
# COLUMNAS AUXILIARES
# =========================
def ensure_columns(ws_master, master_cols: Dict[str, int], names: List[str]) -> Dict[str, int]:
    next_col = ws_master.max_column + 1
    out: Dict[str, int] = {}
    for name in names:
        if name not in master_cols:
            ws_master.cell(1, next_col).value = name
            ws_master.cell(1, next_col).fill = FILL_HEADER
            ws_master.cell(1, next_col).font = Font(bold=True)
            master_cols[name] = next_col
            next_col += 1
        out[name] = master_cols[name]
    return out


def ensure_source_detail_columns(
    ws_master,
    master_cols: Dict[str, int],
    prefix: str,
    fields: List[str],
) -> Dict[str, int]:
    names = [f"{prefix}_{field}" for field in fields]
    return ensure_columns(ws_master, master_cols, names)



def clear_cells(ws_master, row_num: int, cols: Dict[str, int]) -> None:
    for _, col in cols.items():
        cell = ws_master.cell(row_num, col)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)


# =========================
# CONFLICTOS POR FUENTE
# =========================
def apply_conflicts_to_source_columns(
    ws_master,
    row_num: int,
    master_cols: Dict[str, int],
    detail_cols: Dict[str, int],
    source_row: Dict[str, Any],
    compare_fields: List[str],
    prefix: str,
    master_fill,
    ref_fill,
) -> List[str]:
    diff_fields: List[str] = []

    for field in compare_fields:
        if field not in master_cols:
            continue

        master_cell = ws_master.cell(row_num, master_cols[field])
        source_value = source_row.get(field)

        master_cmp = comparable_value(field, master_cell.value)
        source_cmp = comparable_value(field, source_value)

        detail_name = f"{prefix}_{field}"
        if detail_name in detail_cols:
            detail_cell = ws_master.cell(row_num, detail_cols[detail_name])
            detail_cell.value = None
            detail_cell.fill = PatternFill(fill_type=None)

        if master_cmp is not None and source_cmp is not None and master_cmp != source_cmp:
            master_cell.fill = master_fill

            if detail_name in detail_cols:
                detail_cell = ws_master.cell(row_num, detail_cols[detail_name])
                detail_cell.value = source_value
                detail_cell.fill = ref_fill

            diff_fields.append(field)

    return diff_fields


# =========================
# LIMPIEZA COLUMNAS VACIAS
# =========================
def remove_empty_reference_columns(ws):
    prefixes = ("ref_", "jer_", "loccsv_", "jsonref_")
    legacy_cols = {"diff_ref", "diff_jer", "diff_loccsv", "diff_json", "fuentes_con_diferencias"}

    cols_to_delete = []

    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if not header:
            continue

        header = str(header)

        # borrar siempre columnas legacy del diseño viejo
        if header in legacy_cols:
            cols_to_delete.append(col)
            continue

        # borrar columnas auxiliares vacías
        if not any(header.startswith(p) for p in prefixes):
            continue

        has_value = False
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value not in (None, "", " "):
                has_value = True
                break

        if not has_value:
            cols_to_delete.append(col)

    for col in reversed(cols_to_delete):
        ws.delete_cols(col)

# =========================
# LEYENDA / AUDITORIA
# =========================
def crear_leyenda_colores(wb):
    ws = recreate_sheet(wb, SHEET_LEGEND)

    headers = ["Muestra", "Significado", "Fuente", "Qué indica"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(1, col).fill = FILL_HEADER
        ws.cell(1, col).font = Font(bold=True)

    rows = [
        (FILL_FROM_BD, "Dato copiado desde BD", "BD 202603", "El script completó el dato desde la base de sucursales"),
        (FILL_FROM_LOC, "Dato copiado desde Maestro Localidades", "Maestro Localidades - Corregido", "El valor se completó usando cod_localidad o provincia + localidad"),
        (FILL_FROM_JER, "Dato copiado desde jerarquia.csv", "jerarquia.csv", "El valor se completó usando cod_localidad o provincia + localidad"),
        (FILL_FROM_LOCCSV, "Dato copiado desde localidades.csv", "localidades.csv", "El valor se completó como validación secundaria"),
        (FILL_FROM_JSON, "Dato copiado desde JSON zip", "argentina_zip_codes.json", "El valor se completó usando provincia + localidad"),
        (FILL_FROM_WEB, "Dato enriquecido por web", "OpenStreetMap / Nominatim", "El script encontró el dato online"),
        (FILL_NORMALIZED, "Dato normalizado", "Pipeline", "El valor fue corregido de formato"),

        (FILL_CONFLICT_MASTER, "Conflicto en Master", "Cualquier referencia", "La celda original del Master difiere de alguna fuente de referencia"),
        (FILL_CONFLICT_REF, "Valor alternativo ref principal", "ID localidad-provincia-region", "Se muestra en columnas ref_*"),
        (FILL_CONFLICT_JER, "Valor alternativo jerarquia.csv", "jerarquia.csv", "Se muestra en columnas jer_*"),
        (FILL_CONFLICT_LOCCSV, "Valor alternativo localidades.csv", "localidades.csv", "Se muestra en columnas loccsv_*"),
        (FILL_CONFLICT_JSON, "Valor alternativo JSON zip", "argentina_zip_codes.json", "Se muestra en columnas jsonref_*"),

        (PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), "Sin cambios", "-", "El valor ya estaba correcto"),
    ]

    for idx, (fill, significado, fuente, indica) in enumerate(rows, start=2):
        ws.cell(idx, 1).fill = fill
        ws.cell(idx, 2).value = significado
        ws.cell(idx, 3).value = fuente
        ws.cell(idx, 4).value = indica

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 72


def crear_auditoria_geografia(
    wb,
    audit_counter_ref: Dict[str, int],
    audit_counter_jer: Dict[str, int],
    audit_counter_loccsv: Dict[str, int],
    audit_counter_json: Dict[str, int],
    stats: Stats,
):
    ws = recreate_sheet(wb, SHEET_AUDIT)
    headers = ["fuente", "tipo_diferencia", "cantidad"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(1, col).fill = FILL_HEADER
        ws.cell(1, col).font = Font(bold=True)

    ordered_ref_fields = [
        "region_pais_cod", "amba/interior", "region_cod", "region_descr",
        "zona_cod", "zona_descr", "cod_prov", "provincia", "cod_localidad", "localidad"
    ]
    ordered_jer_fields = [
        "region_pais_cod", "amba/interior", "region_cod", "region_descr",
        "zona_cod", "zona_descr", "cod_prov", "provincia", "cod_localidad",
        "localidad", "departamento", "cp"
    ]
    ordered_loccsv_fields = ["cp", "localidad"]
    ordered_json_fields = ["cp", "departamento"]

    for field in ordered_ref_fields:
        ws.append(["ID localidad-provincia-region", field, audit_counter_ref.get(field, 0)])
    ws.append(["ID localidad-provincia-region", "filas_con_conflicto", stats.conflict_rows_ref])
    ws.append(["ID localidad-provincia-region", "conflictos_totales", stats.conflicts_ref])
    ws.append([])

    for field in ordered_jer_fields:
        ws.append(["jerarquia.csv", field, audit_counter_jer.get(field, 0)])
    ws.append(["jerarquia.csv", "filas_con_conflicto", stats.conflict_rows_jer])
    ws.append(["jerarquia.csv", "conflictos_totales", stats.conflicts_jer])
    ws.append([])

    for field in ordered_loccsv_fields:
        ws.append(["localidades.csv", field, audit_counter_loccsv.get(field, 0)])
    ws.append(["localidades.csv", "filas_con_conflicto", stats.conflict_rows_loccsv])
    ws.append(["localidades.csv", "conflictos_totales", stats.conflicts_loccsv])
    ws.append([])

    for field in ordered_json_fields:
        ws.append(["argentina_zip_codes.json", field, audit_counter_json.get(field, 0)])
    ws.append(["argentina_zip_codes.json", "filas_con_conflicto", stats.conflict_rows_json])
    ws.append(["argentina_zip_codes.json", "conflictos_totales", stats.conflicts_json])
    ws.append([])

    ws.append(["completado", "desde_BD", stats.from_bd])
    ws.append(["completado", "desde_Maestro_Localidades", stats.from_loc])
    ws.append(["completado", "desde_jerarquia_csv", stats.from_jer])
    ws.append(["completado", "desde_localidades_csv", stats.from_loccsv])
    ws.append(["completado", "desde_json_zip", stats.from_json])
    ws.append(["completado", "desde_web", stats.from_web])
    ws.append(["pipeline", "celdas_normalizadas", stats.normalized])
    ws.append(["pipeline", "filas_tocadas", stats.rows_touched])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14


# =========================
# MAIN
# =========================
def process_workbook(filename: str, use_web: bool = False, web_limit: int = 20, save_as: Optional[str] = None) -> str:
    filename = str(Path(filename).resolve())
    wb = openpyxl.load_workbook(filename)

    for required in [SHEET_MASTER, SHEET_BD, SHEET_LOC, SHEET_REF]:
        if required not in wb.sheetnames:
            raise ValueError(f"No se encontró la hoja requerida: {required}")

    ws_master = wb[SHEET_MASTER]
    ws_bd = wb[SHEET_BD]
    ws_loc = wb[SHEET_LOC]
    ws_ref = wb[SHEET_REF]

    master_cols = find_headers(ws_master)
    bd_cols = find_headers(ws_bd)
    loc_cols = find_headers(ws_loc)
    ref_cols = find_headers(ws_ref)

    required_master = ["ID Sucursal"] + MASTER_FIELDS
    missing = [c for c in required_master if c not in master_cols]
    if missing:
        raise ValueError(f"Faltan columnas en {SHEET_MASTER}: {missing}")

    if "Id local" not in ref_cols:
        raise ValueError(f"Falta la columna 'Id local' en {SHEET_REF}")

    bd_index = build_bd_index(ws_bd, bd_cols)
    loc_by_cod, loc_by_prov_loc = build_loc_indexes(ws_loc, loc_cols)
    ref_index = build_ref_index(ws_ref, ref_cols)
    jer_by_cod, jer_by_prov_loc = build_jerarquia_indexes(CSV_JERARQUIA)
    loccsv_by_cp_loc = build_localidades_csv_index(CSV_LOCALIDADES)
    json_by_prov_loc, json_by_prov_dept_loc = load_json_zip_indexes(JSON_ZIP)

    detail_ref = ensure_source_detail_columns(ws_master, master_cols, "ref", GEO_FIELDS)
    detail_jer = ensure_source_detail_columns(ws_master, master_cols, "jer", GEO_FIELDS)
    detail_loccsv = ensure_source_detail_columns(ws_master, master_cols, "loccsv", GEO_FIELDS)
    detail_json = ensure_source_detail_columns(ws_master, master_cols, "jsonref", GEO_FIELDS)

    master_cols = find_headers(ws_master)

    stats = Stats()
    web_calls = 0

    audit_counter_ref: Dict[str, int] = {}
    audit_counter_jer: Dict[str, int] = {}
    audit_counter_loccsv: Dict[str, int] = {}
    audit_counter_json: Dict[str, int] = {}

    for row_num in range(2, ws_master.max_row + 1):
        row_touched = normalize_row(ws_master, row_num, master_cols, stats)

        key = normalize_id_sucursal(ws_master.cell(row_num, master_cols["ID Sucursal"]).value)
        cod_loc = normalize_field_value("cod_localidad", ws_master.cell(row_num, master_cols["cod_localidad"]).value)
        prov_key = normalize_key(ws_master.cell(row_num, master_cols["provincia"]).value)
        loc_key = normalize_key(ws_master.cell(row_num, master_cols["localidad"]).value)
        dept_key = normalize_key(ws_master.cell(row_num, master_cols["departamento"]).value)
        cp_norm = parse_int_like(ws_master.cell(row_num, master_cols["cp"]).value)

        # completar desde BD
        bd_row = bd_index.get(key)
        if bd_row:
            for field in MASTER_FIELDS:
                if field in master_cols and set_if_missing(ws_master.cell(row_num, master_cols[field]), bd_row.get(field), FILL_FROM_BD):
                    stats.from_bd += 1
                    row_touched = True

        cod_loc = normalize_field_value("cod_localidad", ws_master.cell(row_num, master_cols["cod_localidad"]).value)
        prov_key = normalize_key(ws_master.cell(row_num, master_cols["provincia"]).value)
        loc_key = normalize_key(ws_master.cell(row_num, master_cols["localidad"]).value)

        # completar desde Maestro Localidades
        loc_row = loc_by_cod.get(cod_loc) if cod_loc else None
        if not loc_row and prov_key and loc_key:
            loc_row = loc_by_prov_loc.get((prov_key, loc_key))

        if loc_row:
            for field in [
                "region_pais_cod", "amba/interior", "region_cod", "region_descr",
                "zona_cod", "zona_descr", "cod_prov", "provincia", "cod_localidad",
                "localidad", "departamento", "cp"
            ]:
                if field in master_cols and set_if_missing(ws_master.cell(row_num, master_cols[field]), loc_row.get(field), FILL_FROM_LOC):
                    stats.from_loc += 1
                    row_touched = True

        cod_loc = normalize_field_value("cod_localidad", ws_master.cell(row_num, master_cols["cod_localidad"]).value)
        prov_key = normalize_key(ws_master.cell(row_num, master_cols["provincia"]).value)
        loc_key = normalize_key(ws_master.cell(row_num, master_cols["localidad"]).value)
        dept_key = normalize_key(ws_master.cell(row_num, master_cols["departamento"]).value)

        # completar desde jerarquia.csv
        jer_row = jer_by_cod.get(cod_loc) if cod_loc else None
        if not jer_row and prov_key and loc_key:
            jer_row = jer_by_prov_loc.get((prov_key, loc_key))

        if jer_row:
            for field in [
                "region_pais_cod", "amba/interior", "region_cod", "region_descr",
                "zona_cod", "zona_descr", "cod_prov", "provincia", "cod_localidad",
                "localidad", "departamento", "cp"
            ]:
                if field in master_cols and set_if_missing(ws_master.cell(row_num, master_cols[field]), jer_row.get(field), FILL_FROM_JER):
                    stats.from_jer += 1
                    row_touched = True

        prov_key = normalize_key(ws_master.cell(row_num, master_cols["provincia"]).value)
        loc_key = normalize_key(ws_master.cell(row_num, master_cols["localidad"]).value)
        dept_key = normalize_key(ws_master.cell(row_num, master_cols["departamento"]).value)

        # completar desde JSON
        json_row = None
        if prov_key and dept_key and loc_key:
            json_row = json_by_prov_dept_loc.get((prov_key, dept_key, loc_key))
        if not json_row and prov_key and loc_key:
            json_row = json_by_prov_loc.get((prov_key, loc_key))

        if json_row:
            for field in ["cp", "departamento"]:
                if field in master_cols and set_if_missing(ws_master.cell(row_num, master_cols[field]), json_row.get(field), FILL_FROM_JSON):
                    stats.from_json += 1
                    row_touched = True

        loc_key = normalize_key(ws_master.cell(row_num, master_cols["localidad"]).value)

        # completar desde localidades.csv
        if loc_key and is_blank(ws_master.cell(row_num, master_cols["cp"]).value):
            matches = [v for (_cp, _loc), v in loccsv_by_cp_loc.items() if _loc == loc_key]
            unique_cps = sorted({m.get("cp") for m in matches if m.get("cp")})
            if len(unique_cps) == 1:
                cp_value = unique_cps[0]
                if set_if_missing(ws_master.cell(row_num, master_cols["cp"]), cp_value, FILL_FROM_LOCCSV):
                    stats.from_loccsv += 1
                    row_touched = True

        # completar desde web
        if use_web and web_calls < web_limit:
            cp_cell = ws_master.cell(row_num, master_cols["cp"])
            dept_cell = ws_master.cell(row_num, master_cols["departamento"])

            if is_blank(cp_cell.value) or is_blank(dept_cell.value):
                addr = ws_master.cell(row_num, master_cols["direccion"]).value
                locality = ws_master.cell(row_num, master_cols["localidad"]).value
                province = ws_master.cell(row_num, master_cols["provincia"]).value

                if not is_blank(addr) and not is_blank(locality):
                    cp, dept = enrich_from_web(str(addr), str(locality), str(province or ""))
                    if is_blank(cp_cell.value) and cp is not None:
                        cp_cell.value = cp
                        cp_cell.fill = FILL_FROM_WEB
                        stats.from_web += 1
                        row_touched = True
                    if is_blank(dept_cell.value) and dept is not None:
                        dept_cell.value = dept
                        dept_cell.fill = FILL_FROM_WEB
                        stats.from_web += 1
                        row_touched = True

                    web_calls += 1
                    time.sleep(1.1)

        if normalize_row(ws_master, row_num, master_cols, stats):
            row_touched = True

        cod_loc = normalize_field_value("cod_localidad", ws_master.cell(row_num, master_cols["cod_localidad"]).value)
        prov_key = normalize_key(ws_master.cell(row_num, master_cols["provincia"]).value)
        loc_key = normalize_key(ws_master.cell(row_num, master_cols["localidad"]).value)
        dept_key = normalize_key(ws_master.cell(row_num, master_cols["departamento"]).value)
        cp_norm = parse_int_like(ws_master.cell(row_num, master_cols["cp"]).value)

        # limpiar columnas compactas y detalles
        clear_cells(ws_master, row_num, detail_ref)
        clear_cells(ws_master, row_num, detail_jer)
        clear_cells(ws_master, row_num, detail_loccsv)
        clear_cells(ws_master, row_num, detail_json)

        # conflictos ref
        ref_row = ref_index.get(key) if key else None
        diff_ref_fields: List[str] = []
        if ref_row:
            ref_row_full = {field: None for field in GEO_FIELDS}
            for field in GEO_FIELDS:
                ref_row_full[field] = ref_row.get(field)
            diff_ref_fields = apply_conflicts_to_source_columns(
                ws_master, row_num, master_cols, detail_ref, ref_row_full,
                GEO_FIELDS, "ref",
                FILL_CONFLICT_MASTER, FILL_CONFLICT_REF
            )
            if diff_ref_fields:
                stats.conflict_rows_ref += 1
                stats.conflicts_ref += len(diff_ref_fields)
                row_touched = True
                for field in diff_ref_fields:
                    audit_counter_ref[field] = audit_counter_ref.get(field, 0) + 1

        # conflictos jer
        jer_row = jer_by_cod.get(cod_loc) if cod_loc else None
        if not jer_row and prov_key and loc_key:
            jer_row = jer_by_prov_loc.get((prov_key, loc_key))

        diff_jer_fields: List[str] = []
        if jer_row:
            jer_row_full = {field: None for field in GEO_FIELDS}
            for field in GEO_FIELDS:
                jer_row_full[field] = jer_row.get(field)
            diff_jer_fields = apply_conflicts_to_source_columns(
                ws_master, row_num, master_cols, detail_jer, jer_row_full,
                GEO_FIELDS, "jer",
                FILL_CONFLICT_MASTER, FILL_CONFLICT_JER
            )
            if diff_jer_fields:
                stats.conflict_rows_jer += 1
                stats.conflicts_jer += len(diff_jer_fields)
                row_touched = True
                for field in diff_jer_fields:
                    audit_counter_jer[field] = audit_counter_jer.get(field, 0) + 1

        # conflictos localidades.csv
        loccsv_row = None
        if cp_norm and loc_key:
            loccsv_row = loccsv_by_cp_loc.get((cp_norm, loc_key))
        if not loccsv_row and loc_key:
            matches = [v for (_cp, _loc), v in loccsv_by_cp_loc.items() if _loc == loc_key]
            if len(matches) == 1:
                loccsv_row = {
                    "cp": matches[0].get("cp"),
                    "localidad": matches[0].get("localidad"),
                }

        diff_loccsv_fields: List[str] = []
        if loccsv_row:
            loccsv_row_full = {field: None for field in GEO_FIELDS}
            loccsv_row_full["cp"] = loccsv_row.get("cp")
            loccsv_row_full["localidad"] = loccsv_row.get("localidad")
            diff_loccsv_fields = apply_conflicts_to_source_columns(
                ws_master, row_num, master_cols, detail_loccsv, loccsv_row_full,
                GEO_FIELDS, "loccsv",
                FILL_CONFLICT_MASTER, FILL_CONFLICT_LOCCSV
            )
            if diff_loccsv_fields:
                stats.conflict_rows_loccsv += 1
                stats.conflicts_loccsv += len(diff_loccsv_fields)
                row_touched = True
                for field in diff_loccsv_fields:
                    audit_counter_loccsv[field] = audit_counter_loccsv.get(field, 0) + 1

        # conflictos json
        json_row = None
        if prov_key and dept_key and loc_key:
            json_row = json_by_prov_dept_loc.get((prov_key, dept_key, loc_key))
        if not json_row and prov_key and loc_key:
            json_row = json_by_prov_loc.get((prov_key, loc_key))

        diff_json_fields: List[str] = []
        if json_row:
            json_row_full = {field: None for field in GEO_FIELDS}
            json_row_full["cp"] = json_row.get("cp")
            json_row_full["departamento"] = json_row.get("departamento")
            diff_json_fields = apply_conflicts_to_source_columns(
                ws_master, row_num, master_cols, detail_json, json_row_full,
                GEO_FIELDS, "jsonref",
                FILL_CONFLICT_MASTER, FILL_CONFLICT_JSON
            )
            if diff_json_fields:
                stats.conflict_rows_json += 1
                stats.conflicts_json += len(diff_json_fields)
                row_touched = True
                for field in diff_json_fields:
                    audit_counter_json[field] = audit_counter_json.get(field, 0) + 1

        if row_touched:
            stats.rows_touched += 1

    crear_leyenda_colores(wb)
    crear_auditoria_geografia(
        wb,
        audit_counter_ref=audit_counter_ref,
        audit_counter_jer=audit_counter_jer,
        audit_counter_loccsv=audit_counter_loccsv,
        audit_counter_json=audit_counter_json,
        stats=stats,
    )

    remove_empty_reference_columns(ws_master)

    output = save_as or filename
    wb.save(output)

    print("\n=== RESUMEN ===")
    print(f"Archivo guardado en: {output}")
    print(f"Celdas completadas desde BD 202603: {stats.from_bd}")
    print(f"Celdas completadas desde Maestro Localidades: {stats.from_loc}")
    print(f"Celdas completadas desde jerarquia.csv: {stats.from_jer}")
    print(f"Celdas completadas desde localidades.csv: {stats.from_loccsv}")
    print(f"Celdas completadas desde argentina_zip_codes.json: {stats.from_json}")
    print(f"Celdas completadas desde web: {stats.from_web}")
    print(f"Celdas normalizadas: {stats.normalized}")
    print(f"Filas tocadas: {stats.rows_touched}")
    print(f"Filas con conflicto REF principal: {stats.conflict_rows_ref}")
    print(f"Conflictos REF principal: {stats.conflicts_ref}")
    print(f"Filas con conflicto jerarquia.csv: {stats.conflict_rows_jer}")
    print(f"Conflictos jerarquia.csv: {stats.conflicts_jer}")
    print(f"Filas con conflicto localidades.csv: {stats.conflict_rows_loccsv}")
    print(f"Conflictos localidades.csv: {stats.conflicts_loccsv}")
    print(f"Filas con conflicto JSON zip: {stats.conflict_rows_json}")
    print(f"Conflictos JSON zip: {stats.conflicts_json}")
    print(f"Hojas generadas/actualizadas: {SHEET_LEGEND}, {SHEET_AUDIT}")

    return output


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Pipeline completo de Master Localizacion con columnas por fuente"
    )
    parser.add_argument("--file", default=FILENAME, help="Ruta al archivo Excel")
    parser.add_argument("--save-as", default=None, help="Guardar resultado en otro archivo")
    parser.add_argument("--backup", action="store_true", help="Genera una copia de seguridad antes de guardar")
    parser.add_argument("--web", action="store_true", help="Hace enriquecimiento web para cp/departamento")
    parser.add_argument("--web-limit", type=int, default=20, help="Máximo de búsquedas web por corrida")
    args = parser.parse_args()

    if args.backup:
        backup = make_backup(args.file)
        print(f"Backup creado: {backup}")

    process_workbook(args.file, use_web=args.web, web_limit=args.web_limit, save_as=args.save_as)


if __name__ == "__main__":
    main()