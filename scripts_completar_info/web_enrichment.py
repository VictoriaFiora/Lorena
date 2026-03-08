import pandas as pd
import requests
import time
import os
import openpyxl
from openpyxl.styles import PatternFill

# Config
FILENAME = "BaseDatos.xlsx"
SHEET_MASTER = "Master Localizacion - Corregida"
USER_AGENT = "LorenaEnrichmentBot/1.0 (vicky@example.com)" # Placeholder, usually needs contact info for OSM
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"

# Blue color for web-enriched fields
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

def get_web_localization(address, locality):
    """
    Usa la API de Nominatim para buscar dirección y localidad.
    """
    params = {
        'q': f"{address}, {locality}, Argentina",
        'format': 'json',
        'addressdetails': 1,
        'limit': 1
    }
    headers = {'User-Agent': USER_AGENT}
    
    try:
        response = requests.get(NOMINATIM_URL, params=params, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if data:
                res = data[0].get('address', {})
                cp = res.get('postcode')
                dept = res.get('county') or res.get('state_district') or res.get('city_district')
                return cp, dept
    except Exception as e:
        print(f"  Error en búsqueda web: {e}")
    
    return None, None

def enrich_master_web(filename):
    print(f"Enriching {SHEET_MASTER} via Web Search...")
    
    wb = openpyxl.load_workbook(filename)
    if SHEET_MASTER not in wb.sheetnames:
        print(f"Error: Hoja {SHEET_MASTER} no encontrada.")
        return

    ws = wb[SHEET_MASTER]
    headers = [c.value for c in ws[1]]
    cols = {h: i + 1 for i, h in enumerate(headers) if h}
    
    # Required columns
    req = ['direccion', 'localidad', 'cp', 'departamento']
    for r in req:
        if r not in cols:
            print(f"Error: Columna '{r}' no encontrada.")
            return

    count = 0
    # Iterar sobre las filas
    for row in ws.iter_rows(min_row=2):
        addr = str(row[cols['direccion']-1].value or "").strip()
        loc = str(row[cols['localidad']-1].value or "").strip()
        
        cp_cell = row[cols['cp']-1]
        dept_cell = row[cols['departamento']-1]
        
        # Solo buscamos si están vacíos
        if (not cp_cell.value or str(cp_cell.value).lower() in ['none', 'nan', '0', '']) or \
           (not dept_cell.value or str(dept_cell.value).lower() in ['none', 'nan', '']):
            
            if addr and loc:
                print(f"  Buscando: {addr}, {loc}...")
                new_cp, new_dept = get_web_localization(addr, loc)
                
                if new_cp or new_dept:
                    if not cp_cell.value and new_cp:
                        cp_cell.value = new_cp
                        cp_cell.fill = BLUE_FILL
                    if not dept_cell.value and new_dept:
                        dept_cell.value = new_dept
                        dept_cell.fill = BLUE_FILL
                    count += 1
                
                # Respetar el Rate Limit de Nominatim (1 req per second)
                time.sleep(1.1)
        
        if count >= 20: # Límite para no bloquearse en una sola ejecución si hay muchos
            print("  Límite de 20 búsquedas alcanzado en esta tanda.")
            break

    if count > 0:
        print(f"  Guardando {filename} con {count} actualizaciones web...")
        wb.save(filename)
    else:
        print("  No se encontraron nuevos datos web.")

if __name__ == "__main__":
    enrich_master_web(FILENAME)
