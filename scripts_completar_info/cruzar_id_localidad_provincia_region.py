
from __future__ import annotations

import argparse
import re
import shutil
import unicodedata
from datetime import datetime
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.styles import PatternFill, Font

FILENAME = "BaseDatos.xlsx"
SHEET_MASTER = "Master Localizacion - Corregida"
SHEET_REF = "ID localidad-provincia-region"

# Colores
FILL_DIFF_MASTER = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_DIFF_REF = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
FILL_HEADER = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

# Mapeo de campos: hoja ref -> hoja master
REF_TO_MASTER = {
    "region_pais_cod": "region_pais_cod",
    "amba/interior": "amba/interior",
    "region_cod": "region_cod",
    "region_descip": "region_descr",
    "zona_cod": "zona_cod",
    "zona_descr": "zona_descr",
    "cod_provinica": "cod_prov",
    "provincia": "provincia",
    "cod_localidad": "cod_localidad",
    "localidad": "localidad",
}


def strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c)
    )


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    return s == "" or s.lower() in {"none", "nan", "null", "faltante", "s/d", "sin dato"}


def parse_int_like(value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    s = clean_spaces(str(value))
    s = s.replace(",", "")
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s


def normalize_id_sucursal(value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    s = clean_spaces(str(value)).lower().replace(" ", "")
    if s.startswith("suc_"):
        raw = s[4:]
    else:
        raw = s
    raw = raw.replace(",", "")
    if re.fullmatch(r"\d+\.0+", raw):
        raw = raw.split(".")[0]
    if re.fullmatch(r"\d+", raw):
        return f"suc_{int(raw)}"
    return f"suc_{raw}"


def normalize_generic(value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    s = clean_spaces(str(value))
    s = strip_accents(s)
    if re.fullmatch(r"\d+(\.0+)?", s.replace(",", "")):
        return parse_int_like(s)
    return s.upper()


def comparable_value(field: str, value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    if field in {
        "region_pais_cod", "region_cod", "zona_cod", "cod_prov", "cod_localidad"
    }:
        return parse_int_like(value)
    return normalize_generic(value)


def find_headers(ws) -> Dict[str, int]:
    return {
        str(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value is not None
    }


def make_backup(path: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.replace(".xlsx", f"_backup_{timestamp}.xlsx")
    shutil.copy2(path, backup)
    return backup


def ensure_diff_columns(ws_master, master_cols: Dict[str, int]) -> Dict[str, int]:
    """
    Crea columnas auxiliares:
      - ref_<campo> para guardar el valor de la hoja de referencia cuando hay diferencia
      - diferencias_id_localidad_provincia_region para resumen por fila
    """
    next_col = ws_master.max_column + 1
    ref_cols: Dict[str, int] = {}

    for master_field in REF_TO_MASTER.values():
        col_name = f"ref_{master_field}"
        if col_name in master_cols:
            ref_cols[master_field] = master_cols[col_name]
            continue
        ws_master.cell(1, next_col).value = col_name
        ws_master.cell(1, next_col).fill = FILL_HEADER
        ws_master.cell(1, next_col).font = Font(bold=True)
        ref_cols[master_field] = next_col
        next_col += 1

    summary_col_name = "diferencias_id_localidad_provincia_region"
    if summary_col_name not in master_cols:
        ws_master.cell(1, next_col).value = summary_col_name
        ws_master.cell(1, next_col).fill = FILL_HEADER
        ws_master.cell(1, next_col).font = Font(bold=True)
        master_cols[summary_col_name] = next_col
        next_col += 1

    # Refrescar cabeceras por si se agregaron nuevas
    master_cols = find_headers(ws_master)
    return {
        **{f"ref_{k}": v for k, v in ref_cols.items()},
        "summary": master_cols[summary_col_name],
    }


def build_ref_index(ws_ref, ref_cols: Dict[str, int]) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = {}
    for row in ws_ref.iter_rows(min_row=2, values_only=True):
        row_map = {h: row[idx - 1] for h, idx in ref_cols.items()}
        key = normalize_id_sucursal(row_map.get("Id local"))
        if not key:
            continue
        cleaned = {}
        for ref_field, master_field in REF_TO_MASTER.items():
            cleaned[master_field] = row_map.get(ref_field)
        index[key] = cleaned
    return index


def process_workbook(filename: str, save_as: Optional[str] = None) -> str:
    wb = openpyxl.load_workbook(filename)

    for required in [SHEET_MASTER, SHEET_REF]:
        if required not in wb.sheetnames:
            raise ValueError(f"No se encontró la hoja requerida: {required}")

    ws_master = wb[SHEET_MASTER]
    ws_ref = wb[SHEET_REF]

    master_cols = find_headers(ws_master)
    ref_cols = find_headers(ws_ref)

    if "ID Sucursal" not in master_cols:
        raise ValueError(f"Falta la columna 'ID Sucursal' en {SHEET_MASTER}")
    if "Id local" not in ref_cols:
        raise ValueError(f"Falta la columna 'Id local' en {SHEET_REF}")

    aux_cols = ensure_diff_columns(ws_master, master_cols)
    master_cols = find_headers(ws_master)
    ref_index = build_ref_index(ws_ref, ref_cols)

    total_rows_with_match = 0
    total_differences = 0

    for row_num in range(2, ws_master.max_row + 1):
        master_id = normalize_id_sucursal(ws_master.cell(row_num, master_cols["ID Sucursal"]).value)
        if not master_id:
            continue

        ref_row = ref_index.get(master_id)
        if not ref_row:
            continue

        total_rows_with_match += 1
        row_diffs = []

        # limpiar columnas auxiliares previas
        for master_field in REF_TO_MASTER.values():
            ws_master.cell(row_num, aux_cols[f"ref_{master_field}"]).value = None
            ws_master.cell(row_num, aux_cols[f"ref_{master_field}"]).fill = PatternFill(fill_type=None)

        ws_master.cell(row_num, aux_cols["summary"]).value = None
        ws_master.cell(row_num, aux_cols["summary"]).fill = PatternFill(fill_type=None)

        for master_field in REF_TO_MASTER.values():
            if master_field not in master_cols:
                continue

            master_cell = ws_master.cell(row_num, master_cols[master_field])
            ref_value_raw = ref_row.get(master_field)

            master_cmp = comparable_value(master_field, master_cell.value)
            ref_cmp = comparable_value(master_field, ref_value_raw)

            # solo marcar si ambos existen y son distintos
            if master_cmp is not None and ref_cmp is not None and master_cmp != ref_cmp:
                ref_cell = ws_master.cell(row_num, aux_cols[f"ref_{master_field}"])
                ref_cell.value = ref_value_raw
                ref_cell.fill = FILL_DIFF_REF

                master_cell.fill = FILL_DIFF_MASTER

                row_diffs.append(
                    f"{master_field}: MASTER={master_cell.value} | REF={ref_value_raw}"
                )
                total_differences += 1

        if row_diffs:
            summary_cell = ws_master.cell(row_num, aux_cols["summary"])
            summary_cell.value = " ; ".join(row_diffs)
            summary_cell.fill = FILL_DIFF_REF

    output = save_as or filename
    wb.save(output)

    print("\n=== RESUMEN ===")
    print(f"Archivo guardado en: {output}")
    print(f"Filas con match contra '{SHEET_REF}': {total_rows_with_match}")
    print(f"Diferencias detectadas: {total_differences}")
    print("Se agregaron columnas ref_* y diferencias_id_localidad_provincia_region en Master Localizacion - Corregida.")

    return output


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Cruza Master Localizacion - Corregida contra ID localidad-provincia-region y marca diferencias."
    )
    parser.add_argument("--file", default=FILENAME, help="Ruta al archivo Excel")
    parser.add_argument("--save-as", default=None, help="Guardar el resultado en otro archivo")
    parser.add_argument("--backup", action="store_true", help="Genera una copia de seguridad antes de guardar")
    args = parser.parse_args()

    if args.backup:
        backup = make_backup(args.file)
        print(f"Backup creado: {backup}")

    process_workbook(args.file, save_as=args.save_as)


if __name__ == "__main__":
    main()
