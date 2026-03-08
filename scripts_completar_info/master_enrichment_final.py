import pandas as pd
import json
import os
import openpyxl
from openpyxl.styles import PatternFill
import web_enrichment # Importamos la lógica web que acabamos de crear

# Config
FILENAME = "BaseDatos.xlsx"
SHEET_MASTER = "Master Localizacion - Corregida"
SHEET_SUC = "Tabla Sucursales Localizacion"
JSON_ZIP = "argentina_zip_codes.json"

# Fills
COLOR_JSON = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
COLOR_SUC  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo

def normalize_text(t):
    if t is None or pd.isna(t): return ""
    return str(t).upper().strip()

def run_master_enrichment():
    print(f"--- Enriqueciendo {SHEET_MASTER} ---")
    
    # 1. Cargar JSON de Códigos Postales
    json_map = {}
    if os.path.exists(JSON_ZIP):
        with open(JSON_ZIP, 'r', encoding='utf-8') as f:
            z_data = json.load(f)
            for prov, depts in z_data.items():
                p_n = normalize_text(prov)
                for dept, loc_parents in depts.items():
                    for lp, sub_locs in loc_parents.items():
                        if isinstance(sub_locs, dict):
                            for sl, cp in sub_locs.items(): 
                                json_map[(p_n, normalize_text(sl))] = (str(cp), dept)
    counts = {'json': 0, 'web': 0}

    # 3. Procesar Hoja Master con openpyxl
    wb = openpyxl.load_workbook(FILENAME)
    ws = wb[SHEET_MASTER]
    headers = [c.value for c in ws[1]]
    cols = {h: i + 1 for i, h in enumerate(headers) if h}
    
    for row in ws.iter_rows(min_row=2):
        prov = normalize_text(row[cols['provincia']-1].value)
        loc = normalize_text(row[cols['localidad']-1].value)
        
        cp_cell = row[cols['cp']-1]
        dept_cell = row[cols['departamento']-1]
        
        is_cp_empty = not cp_cell.value or str(cp_cell.value).lower() in ['none', 'nan', '0', '']
        is_dept_empty = not dept_cell.value or str(dept_cell.value).lower() in ['none', 'nan', '']

        # A. Prioridad 1: JSON
        if is_cp_empty or is_dept_empty:
            res = json_map.get((prov, loc))
            if res:
                if is_cp_empty:
                    cp_cell.value = res[0]
                    cp_cell.fill = COLOR_JSON
                if is_dept_empty:
                    dept_cell.value = res[1]
                    dept_cell.fill = COLOR_JSON
                counts['json'] += 1
                is_cp_empty = False
                is_dept_empty = False

    # Guardar progreso intermedio antes de la web
    print(f"  Resultados: JSON={counts['json']}")
    wb.save(FILENAME)
    
    # C. Prioridad 3: Web (Solo si sigue vacío)
    # Nota: Llamamos a la función web que procesa por tandas
    web_enrichment.enrich_master_web(FILENAME)

if __name__ == "__main__":
    run_master_enrichment()
