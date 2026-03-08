import openpyxl
import os

f = "BaseDatos.xlsx"
s_bd = "BD 202603 - Corregida"
s_m = "Master Localizacion - Corregida"

def diagnose():
    wb = openpyxl.load_workbook(f, data_only=False)
    ws_bd = wb[s_bd]
    ws_m = wb[s_m]

    bd_headers = [c.value for c in ws_bd[1]]
    m_headers = [c.value for c in ws_m[1]]

    print("BD Headers:", bd_headers)
    print("M Headers:", m_headers)

    bd_cols = {h: i + 1 for i, h in enumerate(bd_headers) if h}
    m_cols = {h: i + 1 for i, h in enumerate(m_headers) if h}

    cloc_bd_idx = bd_cols.get("cod_localidad")
    cloc_m_idx = m_cols.get("cod_localidad")
    
    print(f"cod_localidad column in BD: {cloc_bd_idx}")
    print(f"cod_localidad column in Master: {cloc_m_idx}")

    print("\n--- BD Sample (First 5 rows) ---")
    for r in range(2, 7):
        cloc = ws_bd.cell(row=r, column=cloc_bd_idx).value
        # Check a few hierarchy columns
        dept_idx = bd_cols.get("departamento_ccu")
        dept_val = ws_bd.cell(row=r, column=dept_idx).value
        print(f"Row {r}: cloc={cloc} (type={type(cloc)}), formula_dept={dept_val}")

    print("\n--- Master Sample (First 5 rows) ---")
    for r in range(2, 7):
        cloc = ws_m.cell(row=r, column=cloc_m_idx).value
        dept_idx_m = m_cols.get("departamento")
        dept_val_m = ws_m.cell(row=r, column=dept_idx_m).value
        print(f"Row {r}: cloc={cloc} (type={type(cloc)}), dept={dept_val_m}")

    wb.close()

if __name__ == "__main__":
    diagnose()
