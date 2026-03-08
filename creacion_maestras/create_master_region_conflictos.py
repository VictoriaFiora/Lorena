
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

FILENAME = "BaseDatos.xlsx"

SHEET_MASTER = "Master Localizacion - Corregida"
SHEET_LEGEND = "Leyenda_Colores"
SHEET_OUTPUT = "Master Region"

BASE_FIELDS = ["region_pais_cod", "amba/interior", "region_cod", "region_descr"]

SOURCE_PREFIXES = {
    "ref": {
        "label": "ID localidad-provincia-region",
        "fields": [f"ref_{f}" for f in BASE_FIELDS],
    },
    "jer": {
        "label": "jerarquia.csv",
        "fields": [f"jer_{f}" for f in BASE_FIELDS],
    },
    "loccsv": {
        "label": "localidades.csv",
        "fields": [f"loccsv_{f}" for f in BASE_FIELDS],
    },
    "jsonref": {
        "label": "argentina_zip_codes.json",
        "fields": [f"jsonref_{f}" for f in BASE_FIELDS],
    },
}

DEFAULT_HEADER_FILL = PatternFill(start_color="6FA8DC", end_color="6FA8DC", fill_type="solid")
DEFAULT_WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
DEFAULT_CONFLICT_MASTER = PatternFill(start_color="E06666", end_color="E06666", fill_type="solid")
DEFAULT_CONFLICT_REF = PatternFill(start_color="8E7CC3", end_color="8E7CC3", fill_type="solid")
DEFAULT_CONFLICT_JER = PatternFill(start_color="C27BA0", end_color="C27BA0", fill_type="solid")
DEFAULT_CONFLICT_LOCCSV = PatternFill(start_color="76A5AF", end_color="76A5AF", fill_type="solid")
DEFAULT_CONFLICT_JSON = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")

HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)

def normalize(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() == "none":
        return ""
    return s

def clone_fill(fill: PatternFill) -> PatternFill:
    return PatternFill(
        fill_type=fill.fill_type,
        start_color=fill.start_color.rgb or fill.start_color.indexed or fill.start_color.theme,
        end_color=fill.end_color.rgb or fill.end_color.indexed or fill.end_color.theme,
    )

def get_fill_from_legend(ws_legend, fuente_name: str, fallback: PatternFill) -> PatternFill:
    if ws_legend is None:
        return fallback
    headers = {ws_legend.cell(1, c).value: c for c in range(1, ws_legend.max_column + 1)}
    fuente_col = headers.get("Fuente")
    muestra_col = headers.get("Muestra")
    if not fuente_col or not muestra_col:
        return fallback
    for r in range(2, ws_legend.max_row + 1):
        fuente = normalize(ws_legend.cell(r, fuente_col).value)
        if fuente == fuente_name:
            cell = ws_legend.cell(r, muestra_col)
            return PatternFill(
                fill_type=cell.fill.fill_type or fallback.fill_type,
                start_color=cell.fill.start_color.rgb or fallback.start_color.rgb,
                end_color=cell.fill.end_color.rgb or fallback.end_color.rgb,
            )
    return fallback

def sheet_delete_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]

def field_group_has_any_data(ws_master, headers: Dict[str, int], field_names: List[str]) -> bool:
    existing = [f for f in field_names if f in headers]
    if not existing:
        return False
    for row in range(2, ws_master.max_row + 1):
        for f in existing:
            if normalize(ws_master.cell(row, headers[f]).value):
                return True
    return False

def build_output_headers(ws_master, headers: Dict[str, int]) -> Tuple[List[str], List[Tuple[str, str]]]:
    output_headers = list(BASE_FIELDS)
    source_columns: List[Tuple[str, str]] = []  # (prefix, field_name)

    for prefix, meta in SOURCE_PREFIXES.items():
        fields = meta["fields"]
        if field_group_has_any_data(ws_master, headers, fields):
            for f in fields:
                if f in headers:
                    output_headers.append(f)
                    source_columns.append((prefix, f))
    return output_headers, source_columns

def collect_unique_rows(ws_master, headers: Dict[str, int], output_headers: List[str]) -> List[List[str]]:
    seen = set()
    rows_out: List[List[str]] = []

    for row_idx in range(2, ws_master.max_row + 1):
        row_vals = [normalize(ws_master.cell(row_idx, headers[h]).value) if h in headers else "" for h in output_headers]

        if all(v == "" for v in row_vals):
            continue

        # Exigir que exista algo en columnas base
        if all(v == "" for v in row_vals[:len(BASE_FIELDS)]):
            continue

        key = tuple(row_vals)
        if key in seen:
            continue

        seen.add(key)
        rows_out.append(row_vals)

    # orden más estable: base fields
    rows_out.sort(key=lambda vals: tuple(vals[:len(BASE_FIELDS)]))
    return rows_out

def autosize(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

def create_master_region(input_file: str, output_file: Optional[str]) -> str:
    wb = openpyxl.load_workbook(input_file)
    ws_master = wb[SHEET_MASTER]
    ws_legend = wb[SHEET_LEGEND] if SHEET_LEGEND in wb.sheetnames else None

    headers = {ws_master.cell(1, c).value: c for c in range(1, ws_master.max_column + 1)}

    missing_base = [h for h in BASE_FIELDS if h not in headers]
    if missing_base:
        raise ValueError(f"Faltan columnas base en '{SHEET_MASTER}': {missing_base}")

    output_headers, source_columns = build_output_headers(ws_master, headers)

    # Colores desde Leyenda_Colores
    fill_header = get_fill_from_legend(ws_legend, "", DEFAULT_HEADER_FILL)
    fill_base = get_fill_from_legend(ws_legend, "-", DEFAULT_WHITE_FILL)
    fill_master_conflict = get_fill_from_legend(ws_legend, "Cualquier referencia", DEFAULT_CONFLICT_MASTER)

    prefix_fill_map = {
        "ref": get_fill_from_legend(ws_legend, "ID localidad-provincia-region", DEFAULT_CONFLICT_REF),
        "jer": get_fill_from_legend(ws_legend, "jerarquia.csv", DEFAULT_CONFLICT_JER),
        "loccsv": get_fill_from_legend(ws_legend, "localidades.csv", DEFAULT_CONFLICT_LOCCSV),
        "jsonref": get_fill_from_legend(ws_legend, "argentina_zip_codes.json", DEFAULT_CONFLICT_JSON),
    }

    rows_out = collect_unique_rows(ws_master, headers, output_headers)

    sheet_delete_if_exists(wb, SHEET_OUTPUT)
    ws_out = wb.create_sheet(SHEET_OUTPUT)

    # Header
    for col_idx, h in enumerate(output_headers, start=1):
        cell = ws_out.cell(1, col_idx, h)
        cell.fill = clone_fill(fill_header)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    for out_row_idx, values in enumerate(rows_out, start=2):
        value_map = dict(zip(output_headers, values))

        # detectar si existe conflicto base vs alguna fuente presente
        base_has_conflict = False
        for base_field in BASE_FIELDS:
            base_val = normalize(value_map.get(base_field))
            for prefix, meta in SOURCE_PREFIXES.items():
                source_field = f"{prefix}_{base_field}"
                if source_field in value_map:
                    source_val = normalize(value_map.get(source_field))
                    if base_val and source_val and base_val != source_val:
                        base_has_conflict = True
                        break
            if base_has_conflict:
                break

        for col_idx, h in enumerate(output_headers, start=1):
            cell = ws_out.cell(out_row_idx, col_idx, values[col_idx - 1])
            cell.alignment = Alignment(vertical="center")

            if h in BASE_FIELDS:
                cell.fill = clone_fill(fill_master_conflict if base_has_conflict else fill_base)
            else:
                prefix = h.split("_", 1)[0]
                cell.fill = clone_fill(prefix_fill_map.get(prefix, fill_base))

    ws_out.freeze_panes = "A2"
    autosize(ws_out)

    if output_file is None:
        p = Path(input_file)
        output_file = str(p.with_name(f"{p.stem}_master_region.xlsx"))

    wb.save(output_file)
    return output_file

def main():
    parser = argparse.ArgumentParser(description="Crea la hoja 'Master Region' sin modificar 'Master Localizacion - Corregida'.")
    parser.add_argument("--input", default=FILENAME, help="Archivo Excel de entrada")
    parser.add_argument("--output", default=None, help="Archivo Excel de salida")
    args = parser.parse_args()

    out = create_master_region(args.input, args.output)
    print(f"Archivo generado: {out}")

if __name__ == "__main__":
    main()
