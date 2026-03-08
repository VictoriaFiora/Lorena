import openpyxl
from openpyxl.styles import PatternFill
import sys
import os

# Asegurar que los scripts en otras carpetas puedan ser importados
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.append(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "scripts_completar_info"))

import sync_localization

FUENTE = "duplicados corredigos.xlsx"
BASE   = "BaseDatos.xlsx"
COPIA  = "Copia de Original Base de datos _ para cliente 20-02-2026 total clientes.xlsx"
SHEET_BD     = "BD 202603 - Corregida"
SHEET_ELIM   = "Eliminados"

# Light Purple for certified corrections
PURPLE_FILL = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
COLOR_NARANJA = "6" # Excel theme color index for orange in the source file

def normalize_id(val):
    """
    Normaliza IDs de sucursal para asegurar formato "suc_N".
    """
    if val is None: return ""
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    
    if s.isdigit():
        return f"suc_{s}"
    
    if s.lower().startswith("suc_"):
        num = s[4:].strip()
        return f"suc_{num}"
        
    return s

def aplicar_duplicados(filename):
    if not os.path.exists(filename):
        print(f"Error: {filename} no existe.")
        return

    print(f"\n--- Procesando {filename} ---")
    
    # 1. Leer duplicados corregidos
    src = openpyxl.load_workbook(FUENTE)
    ws_src = src.active
    src_headers = [c.value for c in ws_src[1]]
    id_col = src_headers.index("ID Sucursal")
    estado_col = src_headers.index("Estado")

    ids_eliminar = set()
    filas_conservar = {} # sid -> {field: val}

    for row in ws_src.iter_rows(min_row=2):
        sid = normalize_id(row[id_col].value)
        if not sid: continue
        estado = str(row[estado_col].value).strip().upper() if row[estado_col].value else ""

        if estado == "ELIMINADO":
            ids_eliminar.add(sid)
        elif estado == "CONSERVADO":
            corrs = {}
            for cell in row:
                # Verificar si tiene el fondo naranja (correccion manual)
                cc = str(cell.fill.start_color.index) if cell.fill and cell.fill.start_color else ""
                if cc == COLOR_NARANJA:
                    corrs[src_headers[cell.column - 1]] = cell.value
            if corrs:
                filas_conservar[sid] = corrs

    print(f"  Eliminar: {len(ids_eliminar)}, Corregir: {len(filas_conservar)}")

    # 2. Abrir archivo destino
    wb = openpyxl.load_workbook(filename)
    ws_bd = wb[SHEET_BD]
    ws_elim = wb[SHEET_ELIM] if SHEET_ELIM in wb.sheetnames else None

    bd_headers = [c.value for c in ws_bd[1]]
    bd_cols = {h: i+1 for i, h in enumerate(bd_headers) if h}
    id_col_bd = bd_cols.get("ID Sucursal")

    # Mapeo de campos
    FIELD_MAP = {
        "Direccion": "Direccion", "gln": "gln", "cod_localidad": "cod_localidad",
        "localidad": "localidad", "provincia": "provincia", "cod_provinica": "cod_provinica",
        "departamento_ccu": "departamento_ccu", "sucursal_codigo_postal": "sucursal_codigo_postal",
        "zona_descr": "zona_descr", "zona_cod": "zona_cod",
        "region_cod_ccu": "region_cod_ccu", "region_descip_ccu": "region_descip_ccu",
        "amba/interior": "amba/interior", "region_pais_cod": "region_pais_cod"
    }

    rows_a_borrar = []
    correcciones_bd = 0

    # 3. Procesar BD
    for row in ws_bd.iter_rows(min_row=2):
        sid = normalize_id(row[id_col_bd-1].value)
        if not sid: continue

        if sid in ids_eliminar:
            # Mover a eliminados si existe la hoja
            if ws_elim:
                elim_headers = [c.value for c in ws_elim[1]]
                elim_cols = {h: i+1 for i, h in enumerate(elim_headers) if h}
                last_row = ws_elim.max_row + 1
                for h, idx in bd_cols.items():
                    dest_col = elim_cols.get(h) or elim_cols.get(h.lower())
                    if dest_col:
                        ws_elim.cell(row=last_row, column=dest_col).value = row[idx-1].value
            rows_a_borrar.append(row[0].row)

        elif sid in filas_conservar:
            corrs = filas_conservar[sid]
            for src_field, val in corrs.items():
                bd_field = FIELD_MAP.get(src_field, src_field)
                dest_col = bd_cols.get(bd_field)
                if dest_col:
                    cell = ws_bd.cell(row=row[0].row, column=dest_col)
                    cell.value = val
                    cell.fill = PURPLE_FILL
                    correcciones_bd += 1

    # Borrar filas
    for rn in sorted(rows_a_borrar, reverse=True):
        ws_bd.delete_rows(rn)

    print(f"  BD local: {len(rows_a_borrar)} eliminadas, {correcciones_bd} corregidas.")

    # 4. Sincronizar Master Tables (NUEVA LÓGICA)
    # sync_all_masters ahora es inteligente sobre donde poner cada cosa
    sync_localization.sync_all_masters(wb, filas_conservar, is_certified=True)

    print(f"  Guardando {filename}...")
    wb.save(filename)
    wb.close()

if __name__ == "__main__":
    aplicar_duplicados("BaseDatos_Clean.xlsx")
