from __future__ import annotations

import argparse
import math
import os
import re
import shutil
import time
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, Iterable, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill

try:
    import requests  # optional, only used with --web
except Exception:  # pragma: no cover
    requests = None

FILENAME = "BaseDatos.xlsx"
SHEET_MASTER = "Master Localizacion - Corregida"
SHEET_BD = "BD 202603"
SHEET_LOC = "Maestro Localidades - Corregido"

# Color coding
FILL_FROM_BD = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
FILL_FROM_LOC = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
FILL_NORMALIZED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_FROM_WEB = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
USER_AGENT = "MasterLocalizacionCleaner/1.0"

MASTER_FIELDS = [
    "GLN",
    "region_pais_cod",
    "amba/interior",
    "region_cod",
    "region_descr",
    "zona_cod",
    "zona_descr",
    "cod_prov",
    "provincia",
    "cod_localidad",
    "localidad",
    "departamento",
    "cp",
    "direccion",
    "lat",
    "long",
]

BD_TO_MASTER = {
    "gln": "GLN",
    "region_pais_cod": "region_pais_cod",
    "amba/interior": "amba/interior",
    "region_cod_ccu": "region_cod",
    "region_descip_ccu": "region_descr",
    "zona_cod": "zona_cod",
    "zona_descr": "zona_descr",
    "cod_provinica": "cod_prov",
    "provincia": "provincia",
    "cod_localidad": "cod_localidad",
    "localidad": "localidad",
    "departamento_ccu": "departamento",
    "sucursal_codigo_postal": "cp",
    "Direccion": "direccion",
    "latitud": "lat",
    "longitud": "long",
}

LOC_TO_MASTER = {
    "cod_localidad": "cod_localidad",
    "localidad": "localidad",
    "departamento": "departamento",
    "cod_provincia": "cod_prov",
    "provincia": "provincia",
    "cp": "cp",
    "region_pais_cod": "region_pais_cod",
    "amba/interior": "amba/interior",
    "region_cod": "region_cod",
    "region_descr": "region_descr",
    "zona_cod": "zona_cod",
    "zona_descr": "zona_descr",
}

CODE_FIELDS = {"GLN", "region_pais_cod", "region_cod", "zona_cod", "cod_prov", "cod_localidad", "cp"}
TEXT_UPPER_FIELDS = {"amba/interior"}
TITLE_FIELDS = {"provincia", "localidad", "departamento", "region_descr", "zona_descr"}
LATLON_FIELDS = {"lat", "long"}


@dataclass
class Stats:
    from_bd: int = 0
    from_loc: int = 0
    normalized: int = 0
    from_web: int = 0
    rows_touched: int = 0


def strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    return s == "" or s.lower() in {"none", "nan", "null", "faltante", "s/d", "sin dato"}


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(text: Any) -> str:
    if is_blank(text):
        return ""
    s = clean_spaces(str(text))
    s = strip_accents(s).upper()
    return s


def title_case(text: str) -> str:
    small = {"de", "del", "la", "las", "los", "y", "san", "santa"}
    words = clean_spaces(text).lower().split(" ")
    out = []
    for i, w in enumerate(words):
        if i > 0 and w in small:
            out.append(w)
        elif w in {"gba", "caba", "noa", "nea"}:
            out.append(w.upper())
        elif w == "cap":
            out.append("CAP")
        else:
            out.append(w.capitalize())
    return " ".join(out)


def parse_int_like(value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return str(int(value))
        return str(value).replace(",", ".")
    s = clean_spaces(str(value))
    s = s.replace(".0", "") if re.fullmatch(r"\d+\.0", s) else s
    s = s.replace(",", "") if re.fullmatch(r"\d{1,3}(,\d{3})+", s) else s
    s = s.replace(" ", "")
    if re.fullmatch(r"\d+", s):
        return s
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def normalize_id_sucursal(value: Any) -> Optional[str]:
    if is_blank(value):
        return None
    s = clean_spaces(str(value)).lower()
    s = s.replace(" ", "")
    if s.startswith("suc_"):
        raw = s[4:]
    else:
        raw = s
    raw = raw.replace(",", "")
    if re.fullmatch(r"\d+\.0+", raw):
        raw = raw.split(".")[0]
    if re.fullmatch(r"\d+", raw):
        return f"suc_{int(raw)}"
    return f"suc_{raw}"


def normalize_latlon(value: Any) -> Optional[float]:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    s = clean_spaces(str(value))
    # keep decimal comma only when it looks decimal and not thousands
    if re.fullmatch(r"-?\d+,\d+", s):
        s = s.replace(",", ".")
    elif re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", s):
        s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def normalize_field_value(field: str, value: Any) -> Any:
    if is_blank(value):
        return None
    if field == "ID Sucursal":
        return normalize_id_sucursal(value)
    if field in CODE_FIELDS:
        return parse_int_like(value)
    if field in LATLON_FIELDS:
        return normalize_latlon(value)
    s = clean_spaces(str(value))
    if field in TEXT_UPPER_FIELDS:
        return strip_accents(s).upper()
    if field in TITLE_FIELDS:
        return title_case(strip_accents(s))
    return s


def find_headers(ws) -> Dict[str, int]:
    headers = [c.value for c in ws[1]]
    return {str(h): i + 1 for i, h in enumerate(headers) if h is not None}


def build_bd_index(ws_bd, bd_cols: Dict[str, int]) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = {}
    for row in ws_bd.iter_rows(min_row=2, values_only=True):
        row_map = {h: row[idx - 1] for h, idx in bd_cols.items()}
        key = normalize_id_sucursal(row_map.get("ID Sucursal"))
        if not key:
            continue
        cleaned = {}
        for src, dst in BD_TO_MASTER.items():
            cleaned[dst] = normalize_field_value(dst, row_map.get(src))
        # Prefer richer rows if duplicates appear
        score = sum(v is not None for v in cleaned.values())
        prev = index.get(key)
        prev_score = prev.get("__score__", -1) if prev else -1
        if score > prev_score:
            cleaned["__score__"] = score
            index[key] = cleaned
    return index


def build_loc_indexes(ws_loc, loc_cols: Dict[str, int]) -> Tuple[Dict[str, Dict[str, Any]], Dict[Tuple[str, str], Dict[str, Any]]]:
    by_cod: Dict[str, Dict[str, Any]] = {}
    by_prov_loc: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in ws_loc.iter_rows(min_row=2, values_only=True):
        row_map = {h: row[idx - 1] for h, idx in loc_cols.items() if h is not None}
        cleaned = {}
        for src, dst in LOC_TO_MASTER.items():
            if src in row_map:
                cleaned[dst] = normalize_field_value(dst, row_map.get(src))
        cod_loc = cleaned.get("cod_localidad")
        if cod_loc:
            by_cod[cod_loc] = cleaned
        prov = normalize_key(cleaned.get("provincia"))
        loc = normalize_key(cleaned.get("localidad"))
        if prov and loc:
            by_prov_loc[(prov, loc)] = cleaned
    return by_cod, by_prov_loc


def set_if_missing(cell, value, fill) -> bool:
    if value is None:
        return False
    if is_blank(cell.value):
        cell.value = value
        cell.fill = fill
        return True
    return False


def normalize_row(ws, row_num: int, cols: Dict[str, int], stats: Stats) -> bool:
    touched = False
    for field in ["ID Sucursal"] + MASTER_FIELDS:
        if field not in cols:
            continue
        cell = ws.cell(row=row_num, column=cols[field])
        new_val = normalize_field_value(field, cell.value)
        old = cell.value
        comparable_old = None if is_blank(old) else old
        if comparable_old != new_val:
            cell.value = new_val
            if comparable_old is not None or new_val is not None:
                cell.fill = FILL_NORMALIZED
                stats.normalized += 1
                touched = True
    return touched


def enrich_from_web(address: str, locality: str, province: str) -> Tuple[Optional[str], Optional[str]]:
    if requests is None:
        return None, None
    params = {
        "q": f"{address}, {locality}, {province}, Argentina",
        "format": "json",
        "addressdetails": 1,
        "limit": 1,
    }
    headers = {"User-Agent": USER_AGENT}
    try:
        r = requests.get(NOMINATIM_URL, params=params, headers=headers, timeout=15)
        if r.status_code != 200:
            return None, None
        data = r.json()
        if not data:
            return None, None
        addr = data[0].get("address", {})
        cp = addr.get("postcode")
        dept = addr.get("county") or addr.get("state_district") or addr.get("city_district")
        cp = parse_int_like(cp) if cp else None
        dept = title_case(strip_accents(dept)) if dept else None
        return cp, dept
    except Exception:
        return None, None


def process_workbook(filename: str, use_web: bool = False, web_limit: int = 20, save_as: Optional[str] = None) -> str:
    wb = openpyxl.load_workbook(filename)
    for required in [SHEET_MASTER, SHEET_BD, SHEET_LOC]:
        if required not in wb.sheetnames:
            raise ValueError(f"No se encontró la hoja requerida: {required}")

    ws_master = wb[SHEET_MASTER]
    ws_bd = wb[SHEET_BD]
    ws_loc = wb[SHEET_LOC]

    master_cols = find_headers(ws_master)
    bd_cols = find_headers(ws_bd)
    loc_cols = find_headers(ws_loc)

    required_master = ["ID Sucursal"] + MASTER_FIELDS
    missing = [c for c in required_master if c not in master_cols]
    if missing:
        raise ValueError(f"Faltan columnas en {SHEET_MASTER}: {missing}")

    bd_index = build_bd_index(ws_bd, bd_cols)
    loc_by_cod, loc_by_prov_loc = build_loc_indexes(ws_loc, loc_cols)
    stats = Stats()
    web_calls = 0

    for row_num in range(2, ws_master.max_row + 1):
        row_touched = normalize_row(ws_master, row_num, master_cols, stats)
        key = normalize_id_sucursal(ws_master.cell(row_num, master_cols["ID Sucursal"]).value)

        # 1) Complete from BD 202603 by ID Sucursal
        bd_row = bd_index.get(key)
        if bd_row:
            for field in MASTER_FIELDS:
                if field not in master_cols:
                    continue
                if set_if_missing(ws_master.cell(row_num, master_cols[field]), bd_row.get(field), FILL_FROM_BD):
                    stats.from_bd += 1
                    row_touched = True

        # 2) Complete from Maestro Localidades by cod_localidad or provincia+localidad
        cod_loc = normalize_field_value("cod_localidad", ws_master.cell(row_num, master_cols["cod_localidad"]).value)
        prov = normalize_key(ws_master.cell(row_num, master_cols["provincia"]).value)
        loc = normalize_key(ws_master.cell(row_num, master_cols["localidad"]).value)
        loc_row = loc_by_cod.get(cod_loc) if cod_loc else None
        if not loc_row and prov and loc:
            loc_row = loc_by_prov_loc.get((prov, loc))
        if loc_row:
            for field in [
                "region_pais_cod", "amba/interior", "region_cod", "region_descr",
                "zona_cod", "zona_descr", "cod_prov", "provincia", "cod_localidad",
                "localidad", "departamento", "cp"
            ]:
                if field not in master_cols:
                    continue
                if set_if_missing(ws_master.cell(row_num, master_cols[field]), loc_row.get(field), FILL_FROM_LOC):
                    stats.from_loc += 1
                    row_touched = True

        # 3) Optional web fill for missing cp/departamento only
        if use_web and web_calls < web_limit:
            cp_cell = ws_master.cell(row_num, master_cols["cp"])
            dept_cell = ws_master.cell(row_num, master_cols["departamento"])
            if is_blank(cp_cell.value) or is_blank(dept_cell.value):
                addr = ws_master.cell(row_num, master_cols["direccion"]).value
                locality = ws_master.cell(row_num, master_cols["localidad"]).value
                province = ws_master.cell(row_num, master_cols["provincia"]).value
                if not is_blank(addr) and not is_blank(locality):
                    cp, dept = enrich_from_web(str(addr), str(locality), str(province or ""))
                    if is_blank(cp_cell.value) and cp is not None:
                        cp_cell.value = cp
                        cp_cell.fill = FILL_FROM_WEB
                        stats.from_web += 1
                        row_touched = True
                    if is_blank(dept_cell.value) and dept is not None:
                        dept_cell.value = dept
                        dept_cell.fill = FILL_FROM_WEB
                        stats.from_web += 1
                        row_touched = True
                    web_calls += 1
                    time.sleep(1.1)

        # 4) Final pass: normalize again fields we may have just written
        if normalize_row(ws_master, row_num, master_cols, stats):
            row_touched = True

        if row_touched:
            stats.rows_touched += 1

    output = save_as or filename
    wb.save(output)

    print("\n=== RESUMEN ===")
    print(f"Archivo guardado en: {output}")
    print(f"Celdas completadas desde BD 202603: {stats.from_bd}")
    print(f"Celdas completadas desde Maestro Localidades: {stats.from_loc}")
    print(f"Celdas completadas desde web: {stats.from_web}")
    print(f"Celdas normalizadas: {stats.normalized}")
    print(f"Filas tocadas: {stats.rows_touched}")
    return output


def make_backup(path: str) -> str:
    backup = path.replace(".xlsx", "_backup.xlsx")
    shutil.copy2(path, backup)
    return backup


def main() -> None:
    parser = argparse.ArgumentParser(description="Completa y normaliza Master Localizacion - Corregida")
    parser.add_argument("--file", default=FILENAME, help="Ruta al archivo Excel")
    parser.add_argument("--save-as", default=None, help="Guardar resultado en otro archivo")
    parser.add_argument("--backup", action="store_true", help="Genera una copia de seguridad antes de guardar")
    parser.add_argument("--web", action="store_true", help="Hace enriquecimiento web para cp/departamento")
    parser.add_argument("--web-limit", type=int, default=20, help="Máximo de búsquedas web por corrida")
    args = parser.parse_args()

    if args.backup:
        backup = make_backup(args.file)
        print(f"Backup creado: {backup}")

    process_workbook(args.file, use_web=args.web, web_limit=args.web_limit, save_as=args.save_as)


if __name__ == "__main__":
    main()
